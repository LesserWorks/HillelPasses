#!/usr/bin/env python3
import sys
import json
import shutil
import random
import string
import openpyxl
import tempfile
import subprocess
from pathlib import Path

def main():
    num_args = len(sys.argv)
    key_pass = random.choices(string.ascii_lowercase, k=10)
    if num_args == 4:
        pass_dir = sys.argv[1]
        xlsx = sys.argv[2]
        dest = sys.argv[3]
        dest_path = Path("./", dest)
        if Path(dest_path).exists() and Path(dest_path).id_dir():
            shutil.rmtree(dest_path)
        Path(dest_path).mkdir()
        workbook = openpyxl.load_workbook(xlsx)
        worksheet = workbook.active
        prepare_certs(key_pass)
        for row in worksheet.iter_rows(2, worksheet.max_row, values_only=True):
            name = " ".join([row[0], row[1]])
            id = str(row[2])
            barcode = str(row[3])
            makepass(pass_dir, name, id, barcode, Path(dest_path, f"{name.replace(' ', '_')}.pkpass"), key_pass)
        cleanup_certs()

    elif num_args == 5:
        pass_dir = sys.argv[1]
        name = sys.argv[2]
        uid = sys.argv[3]
        barcode = sys.argv[4]
        prepare_certs(key_pass)
        makepass(pass_dir, name, uid, barcode, f"./{uid}.pkpass", key_pass)
        cleanup_certs()
        
    else:
        print("Invalid arguments")
        return
    
def makepass(pass_dir, name, uid, barcode, dest, key_pass):
    if len(uid) != 9:
       print("UID must be 9 digits")
       return
    if len(barcode) != 14:
        print("Barcode must be 14 digits")
        return
    Path(pass_dir, ".DS_Store").unlink(missing_ok=True)
    
    # Make temp directory to make pass in
    with tempfile.TemporaryDirectory() as dst:
        just_folder = Path(pass_dir).name
        newdir = Path(dst, just_folder)
        newdir.mkdir()

        # copy pass source to temp directory
        shutil.copytree(pass_dir, newdir, dirs_exist_ok=True) 
        # open pass.json
        with open(Path(newdir, 'pass.json'), 'r') as pass_file:
            data = json.load(pass_file)
            # write custom data to json object
            data['serialNumber'] = uid
            data['generic']['secondaryFields'][0]['value'] = uid
            data['generic']['primaryFields'][0]['value'] = name
            for code in data['barcodes']:
                code['message'] = barcode
                code['altText'] = barcode

        with open(Path(newdir, 'pass.json'), 'w') as pass_file:
            json.dump(data, pass_file)

        # create manifest file
        manifest = {}
        for file in newdir.iterdir():
            result = subprocess.run(['openssl', 'sha1', file.absolute()], capture_output=True)
            sha1 = result.stdout.decode('ascii').split()[1]
            manifest[file.name] = sha1
        with open(Path(newdir, 'manifest.json'), 'w') as manifest_file:
            json.dump(manifest, manifest_file)
        print(f"Wrote manifest {manifest}")

        # See here for the openssl commands: https://www.kodeco.com/2855-beginning-passbook-in-ios-6-part-1-2/page/4
        subprocess.run(['openssl', 'smime', '-binary', '-sign', '-certfile', 'WWDR.pem',
                    '-signer', 'cert.pem',  '-in', Path(newdir, 'manifest.json'),
                    '-out', Path(newdir, 'signature'), '-outform', 'DER', '-passin', f'pass:{key_pass}'])

        subprocess.run(['zip', '-q', '-j', '-r', dest] + [str(p) for p in newdir.iterdir()])

def prepare_certs(key_pass):
    pass_type_id_cert = 'Certificates.p12'
    subprocess.run(['openssl', 'pkcs12', '-in', pass_type_id_cert, 
                    '-clcerts', '-out', 'cert.pem', '-passin', 'pass:', '-passout', f'pass:{key_pass}'])

def cleanup_certs():
    Path('cert.pem').unlink()

if __name__ == "__main__":
    main()
