#!/usr/bin/env python3
#
# Copyright 2022 Google Inc. All rights reserved.
#
#
# Licensed under the Apache License, Version 2.0 (the "License"); you may not
# use this file except in compliance with the License. You may obtain a copy of
# the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS, WITHOUT
# WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the
# License for the specific language governing permissions and limitations under
# the License.
#
# Extensive modifications by Judah Lesser in 2022
#


import os
import re
import sys
import openpyxl

from google.auth.transport.requests import AuthorizedSession
from google.oauth2 import service_account
from google.auth import jwt, crypt


def main():
    num_args = len(sys.argv)
    if num_args == 3:
        key_file = sys.argv[1]
        xlsx = sys.argv[2]
        
        students = []
        workbook = openpyxl.load_workbook(xlsx)
        worksheet = workbook.active
        for row in worksheet.iter_rows(2, worksheet.max_row, values_only=True):
            name = " ".join([row[0], row[1]])
            id = str(row[2])
            barcode = str(row[3])
            students.append((name, id, barcode)) # read document into tuple
            #print(f"{name} {id} {barcode}")
            #makepass(pass_dir, name, id, barcode, os.path.join(dest_path, f"{name.replace(' ', '')}.pkpass"))
        students.sort(key = lambda x: x[0].replace(" ", "")) # sort students alphabetically
        
        credentials = service_account.Credentials.from_service_account_file(
            key_file,
            scopes=["https://www.googleapis.com/auth/wallet_object.issuer"]
        )
        http_client = AuthorizedSession(credentials)

        for student in students:
            print(student[0]) # print students' names in the order we will generate the links
        for student in students:
            makepasslink(http_client, key_file, student[0], student[1], student[2])

    elif num_args == 5:
        key_file = sys.argv[1]
        name = sys.argv[2]
        uid = sys.argv[3]
        barcode = sys.argv[4]

        credentials = service_account.Credentials.from_service_account_file(
            key_file,
            scopes=["https://www.googleapis.com/auth/wallet_object.issuer"]
        )
        http_client = AuthorizedSession(credentials)

        makepasslink(http_client, key_file, name, uid, barcode)
        
    else:
        print("Invalid arguments")
        return

def makepasslink(http_client, key_file, name, uid, barcode):
    # Make sure these variables are up to date!
    issuer_id = 3388000000022144163
    class_id = "hillelcard"

    if len(uid) != 9:
        print("UID must be 9 digits")
        return
    if len(barcode) != 14:
        print("Barcode must be 14 digits")
        return

    user_id = barcode # This uniquely identifies the pass
    object_id = "%s.%s-%s" % (issuer_id, re.sub(r"[^\w.-]", "_", user_id), class_id)
    object_url = "https://walletobjects.googleapis.com/walletobjects/v1/loyaltyObject/"
    object_payload = {
        "id": object_id,
        "classId": f"{issuer_id}.{class_id}",
        "barcode": {
            "kind": "walletobjects#barcode",
            "type": "CODE_128",
            "value": str(barcode),
            "alternateText": str(barcode)
        },
        "state": "active",
        "accountId": str(uid),
        "accountName": name,
    }

    # Generate pass on Google's server
    object_response = http_client.get(object_url + object_id)
    if object_response.status_code == 404:
        # Object does not yet exist
        # Send POST request to create it
        object_response = http_client.post(
            object_url,
            json=object_payload
        )
    
    # Generate pass access link
    claims = {
        "iss": http_client.credentials.service_account_email,
        "aud": "google",
        "origins": ["www.hilleltournament.com"],
        "typ": "savetowallet",
        "payload": {
            "loyaltyObjects": [
                {
                    "id": object_id
                }
            ]
        }
    }

    signer = crypt.RSASigner.from_service_account_file(key_file)
    token = jwt.encode(signer, claims).decode("utf-8")
    save_url = f"https://pay.google.com/gp/v/save/{token}"

    print(save_url)

if __name__ == "__main__":
    main()
